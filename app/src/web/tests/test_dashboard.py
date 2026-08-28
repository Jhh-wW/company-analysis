"""승인된 운영 대시보드의 권한·차단·HTML 새로고침 계약."""

from __future__ import annotations

import hashlib
from dataclasses import replace
from datetime import date, datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.features.admin_dashboard import kpi as dashboard_kpi
from src.features.admin_dashboard import maintenance as dashboard_maintenance
from src.features.admin_dashboard import store as dashboard_store
from src.features.auth import constants as auth_constants
from src.features.auth import logic as auth_logic
from src.features.backup import status as backup_status
from src.features.pipeline.canonical_demo import build_demo_report
from src.features.pipeline.demo import DemoPipeline
from src.features.report_access import store as report_access_store
from src.features.sharelink import allowlist
from src.features.sharelink import store as share_store
from src.features.storage import constants as storage_constants
from src.features.storage import db, reports
from src.web import main, report_retention_adapter, runtime
from src.web import deployment_mode
from src.web.routers import dashboard as dashboard_router
from src.web.routers import maintenance as maintenance_router
from src.web.routers import reports as reports_router


def _identity_subject(email: str) -> str:
    digest = hashlib.sha256(email.lower().encode("utf-8")).hexdigest()[:24]
    return f"google:test-{digest}"


def _session(client: TestClient, *, email: str, is_admin: bool) -> str:
    session = auth_logic.create_session(
        email,
        is_admin,
        subject=_identity_subject(email),
    )
    client.cookies.set(auth_constants.SESSION_COOKIE_NAME, session.token)
    return auth_logic.csrf_token_for_session(session.token)


def _seed_report() -> str:
    report_id = "da" * 16
    with db.connect() as conn:
        reports.save(conn, report_id, "CORP-001", "", build_demo_report())
    return report_id


def _bind_member_report(conn, *, report_id: str, email: str) -> None:
    """실제 MEMBER 출고처럼 run 예약·불변 계정·report를 한 소유권으로 묶는다."""

    assert dashboard_store.reserve_member_run(
        conn,
        run_id=report_id,
        actor_email=email,
        day="2026-08-22",
        now_iso="2026-08-22T10:00:00+09:00",
    )
    assert report_access_store.bind_member_run(
        conn,
        run_id=report_id,
        identity_subject=_identity_subject(email),
    )
    assert report_access_store.bind_report(
        conn,
        run_id=report_id,
        report_id=report_id,
    )
    assert dashboard_store.settle_member_run(
        conn,
        run_id=report_id,
        succeeded=True,
        report_id=report_id,
        now_iso="2026-08-22T10:01:00+09:00",
    )


def test_admin_dashboard_has_no_public_json_and_refresh_is_admin_only(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    with TestClient(main.app) as client:
        denied = client.get("/admin/refresh/today", follow_redirects=False)
        assert denied.status_code in (302, 303, 307)
        csrf = _session(client, email="admin@example.com", is_admin=True)
        response = client.get("/admin")
        fragment = client.get("/admin/refresh/today")
        settings = client.get("/admin/settings")

    assert csrf
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/html")
    assert "오늘" in response.text and "문제·보고서" in response.text
    assert fragment.status_code == 200 and fragment.headers["content-type"].startswith("text/html")
    assert settings.status_code == 200


def test_member_error_closes_result_and_pdf_until_manual_republish(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with db.connect() as conn:
        assert allowlist.invite(conn, email="member@example.com", note="", now_iso="2026-08-22T10:00:00+09:00")
        _bind_member_report(conn, report_id=report_id, email="member@example.com")
    with TestClient(main.app) as client:
        csrf = _session(client, email="member@example.com", is_admin=False)
        reported = client.post(
            f"/reports/{report_id}/errors", follow_redirects=False,
            data={"area": "매출 표", "reason": "원출처와 다릅니다", "csrf_token": csrf},
        )
        result = client.get(f"/result/{report_id}", follow_redirects=False)
        pdf = client.get(f"/download/pdf/{report_id}", follow_redirects=False)

    assert reported.status_code == 303
    assert result.status_code == 409
    assert pdf.status_code == 409
    assert "오류 신고가 접수되어" in result.text


def test_member_cannot_escalate_own_report_to_a_global_incident(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with db.connect() as conn:
        allowlist.invite(
            conn, email="member@example.com", note="", now_iso="2026-08-22T10:00:00+09:00"
        )
        _bind_member_report(conn, report_id=report_id, email="member@example.com")
    with TestClient(main.app) as member:
        csrf = _session(member, email="member@example.com", is_admin=False)
        response = member.post(
            f"/reports/{report_id}/errors",
            data={
                "area": "표",
                "reason": "수치가 다릅니다",
                "incident_kind": "security",
                "csrf_token": csrf,
            },
            follow_redirects=False,
        )
    with db.connect() as conn:
        service = dashboard_store.get_service_state(conn)
        incidents = dashboard_store.list_incidents(conn)

    assert response.status_code == 303
    assert service.status == dashboard_store.SERVICE_NORMAL
    assert incidents == []


def test_survey_is_member_only_and_revision_is_visible_to_admin(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with db.connect() as conn:
        allowlist.invite(conn, email="member@example.com", note="친구", now_iso="2026-08-22T10:00:00+09:00")
        _bind_member_report(conn, report_id=report_id, email="member@example.com")
    with TestClient(main.app) as member:
        csrf = _session(member, email="member@example.com", is_admin=False)
        viewed = member.get(f"/result/{report_id}")
        saved = member.post(
            f"/reports/{report_id}/survey", follow_redirects=False,
            data={"rating": "5", "overall_feedback": "유용합니다", "business_distinction": "구분점이 잘 보입니다", "csrf_token": csrf},
        )
    with TestClient(main.app) as admin:
        _session(admin, email="admin@example.com", is_admin=True)
        detail = admin.get(f"/admin/reports/{report_id}")

    assert saved.status_code == 303
    assert viewed.status_code == 200
    assert detail.status_code == 200
    assert "member@example.com" in detail.text
    assert "5점" in detail.text
    with db.connect() as conn:
        snapshot = dashboard_store.get_report_snapshot(
            conn, report_id=report_id, version=1
        )
        assert snapshot is not None and len(snapshot.payload_sha256) == 64
        # 공개 GET은 권한 판정뿐 아니라 KPI 사건도 쓰지 않는다. 첫 열람을
        # GET 부작용으로 재도입하지 않는 한 응답시간 KPI는 측정 불가다.
        assert dashboard_kpi.summary(conn) == dashboard_kpi.KpiSummary(
            measured_responses=0,
            within_target=0,
        )


def test_member_survey_prefills_saved_answer_and_keeps_browser_draft(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with db.connect() as conn:
        report = reports.load(conn, report_id)
        assert report is not None
        dashboard_store.register_report(
            conn,
            report_id=report_id,
            corp_type=report.corp_type,
            payload_json=reports.report_to_json(report),
            now_iso="2026-08-22T09:59:00+09:00",
        )
        allowlist.invite(
            conn, email="member@example.com", display_name="김민지", note="스터디",
            now_iso="2026-08-22T10:00:00+09:00",
        )
        _bind_member_report(conn, report_id=report_id, email="member@example.com")
        dashboard_store.save_survey(
            conn, report_id=report_id, actor_email="member@example.com", rating=4,
            overall_feedback="기존 소감", business_distinction="기존 구분점",
            add_information="추가 정보", delete_information="중복 정보",
            now_iso="2026-08-22T10:01:00+09:00",
        )
    with TestClient(main.app) as member:
        _session(member, email="member@example.com", is_admin=False)
        result = member.get(f"/result/{report_id}")
    with TestClient(main.app) as admin:
        _session(admin, email="admin@example.com", is_admin=True)
        members = admin.get("/admin/members")

    assert result.status_code == 200
    assert '<option value="4" selected>4점</option>' in result.text
    assert ">기존 소감</textarea>" in result.text
    assert "window.sessionStorage" in result.text
    assert "설문 수정 저장" in result.text
    assert "김민지 · member@example.com" in members.text
    assert "기존 소감" in members.text and "기존 구분점" in members.text
    assert f'href="/admin/reports/{report_id}/versions/1"' in members.text


def test_member_survey_prefill_and_admin_snapshot_are_scoped_to_report_version(
    monkeypatch, tmp_path
):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    version_one = build_demo_report()
    version_two = replace(
        version_one,
        generated_at="2026-08-23",
    )
    with db.connect() as conn:
        dashboard_store.register_report(
            conn,
            report_id=report_id,
            corp_type=version_one.corp_type,
            payload_json=reports.report_to_json(version_one),
            now_iso="2026-08-22T09:00:00+09:00",
        )
        allowlist.invite(
            conn,
            email="member@example.com",
            display_name="김민지",
            note="스터디",
            now_iso="2026-08-22T09:01:00+09:00",
        )
        _bind_member_report(conn, report_id=report_id, email="member@example.com")
        assert dashboard_store.save_survey(
            conn,
            report_id=report_id,
            report_version=1,
            actor_email="member@example.com",
            rating=5,
            overall_feedback="버전1에서 남긴 소감",
            business_distinction="버전1 구분점",
            add_information="",
            delete_information="",
            now_iso="2026-08-22T10:00:00+09:00",
        ) == 1
        dashboard_store.record_error(
            conn,
            report_id=report_id,
            actor_email="member@example.com",
            area="출처",
            reason="버전2 수정 필요",
            now_iso="2026-08-22T10:01:00+09:00",
        )
        for status in (
            dashboard_store.REPORT_STATUS_FIXING,
            dashboard_store.REPORT_STATUS_RECHECKING,
        ):
            dashboard_store.change_report_state(
                conn,
                report_id=report_id,
                next_status=status,
                actor_email="admin@example.com",
                reason="수정본 재검사",
                now_iso="2026-08-22T10:02:00+09:00",
            )
        assert dashboard_store.capture_report_snapshot(
            conn,
            report_id=report_id,
            version=2,
            payload_json=reports.report_to_json(version_two),
            actor="admin@example.com",
            now_iso="2026-08-22T10:03:00+09:00",
        )
        dashboard_store.change_report_state(
            conn,
            report_id=report_id,
            next_status=dashboard_store.REPORT_STATUS_NORMAL,
            actor_email="admin@example.com",
            reason="수정본 재검사 완료",
            now_iso="2026-08-22T10:04:00+09:00",
        )

    with TestClient(main.app) as member:
        csrf = _session(member, email="member@example.com", is_admin=False)
        current_before = member.get(f"/result/{report_id}")
        saved_version_two = member.post(
            f"/reports/{report_id}/survey",
            data={
                "rating": "3",
                "overall_feedback": "버전2에서 남긴 소감",
                "business_distinction": "버전2 구분점",
                "csrf_token": csrf,
            },
            follow_redirects=False,
        )
        denied_snapshot = member.get(
            f"/admin/reports/{report_id}/versions/1", follow_redirects=False
        )
    with TestClient(main.app) as admin:
        _session(admin, email="admin@example.com", is_admin=True)
        members = admin.get("/admin/members")
        snapshot_one = admin.get(f"/admin/reports/{report_id}/versions/1")
        snapshot_two = admin.get(f"/admin/reports/{report_id}/versions/2")

    assert current_before.status_code == 200
    assert "버전1에서 남긴 소감" not in current_before.text
    assert "설문 저장" in current_before.text
    assert f'data-draft-key="member-survey:{report_id}:v2:' in current_before.text
    assert saved_version_two.status_code == 303
    assert denied_snapshot.status_code in (302, 303, 307)
    assert members.status_code == 200
    assert f'/admin/reports/{report_id}/versions/1' in members.text
    assert f'/admin/reports/{report_id}/versions/2' in members.text
    assert snapshot_one.status_code == 200
    assert "버전1에서 남긴 소감" not in snapshot_one.text
    assert "2026-08-23" not in snapshot_one.text
    assert "과거 스냅샷 · 읽기 전용" in snapshot_one.text
    assert f'action="/admin/reports/{report_id}' not in snapshot_one.text
    assert "수동 상태 변경" not in snapshot_one.text
    assert snapshot_two.status_code == 200
    assert "2026-08-23" in snapshot_two.text


def test_admin_never_falls_back_to_current_report_when_survey_snapshot_sha_is_wrong(
    monkeypatch, tmp_path
):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with db.connect() as conn:
        report = reports.load(conn, report_id)
        assert report is not None
        dashboard_store.register_report(
            conn,
            report_id=report_id,
            corp_type=report.corp_type,
            payload_json=reports.report_to_json(report),
            now_iso="2026-08-22T09:00:00+09:00",
        )
        dashboard_store.save_survey(
            conn,
            report_id=report_id,
            report_version=1,
            actor_email="member@example.com",
            rating=4,
            overall_feedback="검증할 소감",
            business_distinction="검증할 구분점",
            add_information="",
            delete_information="",
            now_iso="2026-08-22T10:00:00+09:00",
        )
        conn.execute(
            f"DROP TRIGGER {dashboard_store.TABLE_REPORT_VERSIONS}_no_update"
        )
        conn.execute(
            f"UPDATE {dashboard_store.TABLE_REPORT_VERSIONS} "
            "SET payload_json = ? WHERE report_id = ? AND version = 1",
            ('{"company":"손상된 현재 대체본"}', report_id),
        )

    with TestClient(main.app) as admin:
        _session(admin, email="admin@example.com", is_admin=True)
        members = admin.get("/admin/members")
        snapshot = admin.get(f"/admin/reports/{report_id}/versions/1")

    assert members.status_code == 200
    assert "당시 보고서 확인 불가" in members.text
    assert f'/admin/reports/{report_id}/versions/1' not in members.text
    assert "손상된 현재 대체본" not in members.text
    assert snapshot.status_code == 503
    assert "무결성을 확인할 수 없습니다" in snapshot.text
    assert "손상된 현재 대체본" not in snapshot.text


def test_dashboard_keeps_working_when_only_survey_summary_is_unavailable(
    monkeypatch, tmp_path
):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()

    def fail_summary(_conn, **_kwargs):
        raise RuntimeError("survey unavailable")

    monkeypatch.setattr(dashboard_router.dashboard_store, "survey_summary", fail_summary)
    with TestClient(main.app) as client:
        _session(client, email="admin@example.com", is_admin=True)
        response = client.get("/admin")

    assert response.status_code == 200
    assert "운영 중" in response.text
    assert "확인 불가" in response.text


def test_issue_priority_and_partial_read_failure_remain_visible(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with db.connect() as conn:
        report = reports.load(conn, report_id)
        assert report is not None
        dashboard_store.register_report(
            conn,
            report_id=report_id,
            corp_type=report.corp_type,
            payload_json=reports.report_to_json(report),
            now_iso="2026-08-22T10:00:00+09:00",
        )
        dashboard_store.record_error(
            conn,
            report_id=report_id,
            actor_email="member@example.com",
            area="개별 신고 문제",
            reason="원출처와 다릅니다",
            now_iso="2026-08-22T10:01:00+09:00",
        )

    warning = {
        "id": 1,
        "kind": "rate_limit",
        "summary": "지연 비용 주의",
        "created_at": "2026-08-22T10:03:00+09:00",
        "report_id": "",
    }
    operation = {
        "operation_key": "weekly:2026-08-17",
        "operation": "weekly_xlsx",
        "detail": "정기 작업 문제",
        "created_at": "2026-08-22T10:02:00+09:00",
    }
    monkeypatch.setattr(
        dashboard_router.dashboard_store,
        "list_active_incidents",
        lambda _conn, **_kwargs: [warning],
    )
    monkeypatch.setattr(
        dashboard_router.dashboard_store,
        "list_failed_operation_issues",
        lambda _conn, **_kwargs: [operation],
    )

    with TestClient(main.app) as client:
        _session(client, email="admin@example.com", is_admin=True)
        ordered = client.get("/admin/issues")

        def fail_incidents(_conn, **_kwargs):
            raise RuntimeError("incidents unavailable")

        monkeypatch.setattr(
            dashboard_router.dashboard_store,
            "list_active_incidents",
            fail_incidents,
        )
        partial = client.get("/admin/issues")

    assert ordered.status_code == 200
    assert ordered.text.index("개별 신고 문제") < ordered.text.index("정기 작업 문제")
    assert ordered.text.index("정기 작업 문제") < ordered.text.index("지연 비용 주의")
    assert partial.status_code == 200
    assert "일부 문제 목록을 확인할 수 없습니다" in partial.text
    assert "개별 신고 문제" in partial.text


def test_narrow_free_admin_demo_disables_deferred_actions(monkeypatch, tmp_path):
    """배포 범위 배너는 내부 사정이라 화면에서 뺐다(2-6) — 여기서는 실제로
    LINK 발급·친구 초대가 막히는지(폼이 없고 버튼이 비활성인지)만 지킨다."""
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    monkeypatch.setenv(
        deployment_mode.ENV_DEPLOYMENT_RUNTIME_CONTRACT,
        deployment_mode.RENDER_ADMIN_DEMO_NO_FORWARDED_CONTRACT,
    )
    monkeypatch.setenv(deployment_mode.ENV_PUBLIC_ORIGIN, "https://demo.example")
    monkeypatch.setenv(auth_constants.ENV_BETA_ADMIN_ONLY, "1")
    runtime._PIPELINE = DemoPipeline()
    with TestClient(main.app, base_url="https://demo.example") as client:
        _session(client, email="admin@example.com", is_admin=True)
        dashboard = client.get("/admin")
        access = client.get("/admin/access")

    assert dashboard.status_code == 200 and access.status_code == 200
    assert "LINK 발급 불가" in access.text and "친구 초대 불가" in access.text
    assert 'action="/admin/link/new"' not in access.text
    assert 'action="/admin/invite"' not in access.text


def test_admin_real_contract_is_admin_only_and_disables_deferred_actions(
    monkeypatch, tmp_path
):
    """배포 범위 배너는 내부 사정이라 화면에서 뺐다(2-6) — 여기서는 관리자 전용
    접근 제어와, 실제로 LINK 발급·친구 초대가 막히는지만 지킨다."""
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    monkeypatch.setenv(
        deployment_mode.ENV_DEPLOYMENT_RUNTIME_CONTRACT,
        deployment_mode.RENDER_ADMIN_REAL_NO_FORWARDED_CONTRACT,
    )
    monkeypatch.setenv(deployment_mode.ENV_PUBLIC_ORIGIN, "https://pilot.example")
    monkeypatch.delenv(auth_constants.ENV_BETA_ADMIN_ONLY, raising=False)
    runtime._PIPELINE = DemoPipeline()

    with TestClient(main.app, base_url="https://pilot.example") as client:
        denied = client.get("/", follow_redirects=False)
        _session(client, email="admin@example.com", is_admin=True)
        dashboard = client.get("/admin")
        access = client.get("/admin/access")

    assert denied.status_code == 303
    assert denied.headers["location"] == "/auth/login"
    assert dashboard.status_code == 200 and access.status_code == 200
    assert "LINK 발급 불가" in access.text and "친구 초대 불가" in access.text
    assert 'action="/admin/link/new"' not in access.text
    assert 'action="/admin/invite"' not in access.text


def test_admin_dashboard_labels_three_minute_metric_as_response_not_accuracy(
    monkeypatch, tmp_path
):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()

    with TestClient(main.app) as client:
        _session(client, email="admin@example.com", is_admin=True)
        response = client.get("/admin/members")

    assert response.status_code == 200
    assert "3분 내 구분점 응답" in response.text
    assert "근거 정확성은 관리자 검토" in response.text


def test_member_revocation_blocks_existing_session_result_and_pdf(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with db.connect() as conn:
        assert allowlist.invite(conn, email="member@example.com", note="", now_iso="2026-08-22T10:00:00+09:00")
        _bind_member_report(conn, report_id=report_id, email="member@example.com")
    with TestClient(main.app) as client:
        _session(client, email="member@example.com", is_admin=False)
        with db.connect() as conn:
            assert allowlist.revoke(conn, "member@example.com")
        result = client.get(f"/result/{report_id}", follow_redirects=False)
        pdf = client.get(f"/download/pdf/{report_id}", follow_redirects=False)

    assert result.status_code == 403
    assert pdf.status_code == 403


def test_link_detail_preserves_history_and_offers_csrf_auto_confirmation(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    raw_key = "a1b2c3d4e5f60718a1b2c3d4e5f60718"
    with db.connect() as conn:
        assert share_store.insert_new(conn, key=raw_key, company="CORP", job="", now_iso="2026-08-22T10:00:00+09:00")
        assert share_store.mark_opened(conn, raw_key, "2026-08-22T10:01:00+09:00")
    key_hash = share_store.key_hash_of(raw_key)
    with TestClient(main.app) as client:
        csrf = _session(client, email="admin@example.com", is_admin=True)
        first = client.get(f"/admin/links/{key_hash}")
        second = client.get(f"/admin/links/{key_hash}")
        confirmed = client.post(
            f"/admin/links/{key_hash}/opens/confirm",
            data={"csrf_token": csrf},
            follow_redirects=False,
        )
        cleared = client.get(f"/admin/links/{key_hash}")

    assert "새 접속 확인" in first.text
    assert "새 접속 확인" in second.text
    assert 'id="link-open-confirm-form"' in first.text
    assert ".requestSubmit()" in first.text
    assert confirmed.status_code == 303
    assert "새로 확인할 접속이 없습니다" in cleared.text
    assert "2026-08-22 10:01 (한국시간) · 누적 1회" in cleared.text


def test_link_list_sorts_by_recent_open_and_marks_unseen(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    with db.connect() as conn:
        assert share_store.insert_new(conn, key="old-link-key", company="OLD", job="", now_iso="2026-08-22T10:00:00+09:00")
        assert share_store.insert_new(conn, key="new-link-key", company="NEW", job="", now_iso="2026-08-22T10:01:00+09:00")
        assert share_store.mark_opened(conn, "old-link-key", "2026-08-22T10:02:00+09:00")
        assert share_store.mark_opened(conn, "new-link-key", "2026-08-22T10:03:00+09:00")
    with TestClient(main.app) as client:
        _session(client, email="admin@example.com", is_admin=True)
        response = client.get("/admin/links")

    assert response.status_code == 200
    assert response.text.index("NEW") < response.text.index("OLD")
    assert "새 접속 1건" in response.text


def test_same_window_aggregate_reports_only_new_count_after_confirmation(
    monkeypatch,
    tmp_path,
):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    raw_key = "aggregated-link-open-secret"
    opened_at = "2026-08-23T10:00:30+09:00"
    with db.connect() as conn:
        assert share_store.insert_new(
            conn,
            key=raw_key,
            company="AGGREGATE",
            job="",
            now_iso="2026-08-23T09:00:00+09:00",
        )
        assert share_store.mark_opened(conn, raw_key, opened_at)
        assert share_store.mark_opened(conn, raw_key, opened_at)
    key_hash = share_store.key_hash_of(raw_key)

    with TestClient(main.app) as client:
        csrf = _session(client, email="admin@example.com", is_admin=True)
        first_list = client.get("/admin/links")
        first_detail = client.get(f"/admin/links/{key_hash}")
        confirmed = client.post(
            f"/admin/links/{key_hash}/opens/confirm",
            data={"csrf_token": csrf},
            follow_redirects=False,
        )
        with db.connect() as conn:
            assert share_store.mark_opened(conn, raw_key, opened_at)
        second_list = client.get("/admin/links")
        second_detail = client.get(f"/admin/links/{key_hash}")

    assert "새 접속 2건" in first_list.text
    assert "2026-08-23 10:00 (한국시간) · 누적 2회" in first_detail.text
    assert confirmed.status_code == 303
    assert "새 접속 1건" in second_list.text
    assert "2026-08-23 10:00 (한국시간) · 누적 3회" in second_detail.text
    assert "새 접속 1회" in second_detail.text


def test_member_page_has_period_controls_and_does_not_fake_legacy_summary(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    with TestClient(main.app) as client:
        _session(client, email="admin@example.com", is_admin=True)
        response = client.get("/admin/members?period=7d")

    assert response.status_code == 200
    assert "최근 7일" in response.text
    assert "확인 불가" in response.text


def test_settings_default_is_read_only_and_component_states_are_not_faked(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    monkeypatch.setenv("GOOGLE_CLIENT_ID", "fixture-client-id")
    monkeypatch.delenv("GOOGLE_API_KEY", raising=False)
    runtime._PIPELINE = DemoPipeline()
    with TestClient(main.app) as client:
        _session(client, email="admin@example.com", is_admin=True)
        readonly = client.get("/admin/settings")
        change = client.get("/admin/settings/change")

    assert readonly.status_code == 200
    assert "변경하기" in readonly.text
    assert 'action="/admin/settings/service"' not in readonly.text
    assert "아직 사용 안 함" in readonly.text
    assert '<h2>Google</h2><p><strong>확인 불가</strong>' in readonly.text
    assert "이미 완료됐거나 다른 실행이 진행 중이면" in change.text
    assert "오늘 정리가 이미 완료됐거나 진행 중이면" in change.text
    assert change.status_code == 200
    assert 'action="/admin/settings/service"' in change.text


def test_settings_shows_persisted_backup_failure_instead_of_configuration_guess(
    monkeypatch, tmp_path
):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    monkeypatch.setattr(
        dashboard_router.clock,
        "iso_now_kst",
        lambda: "2026-08-22T12:00:00+09:00",
    )
    runtime._PIPELINE = DemoPipeline()
    with db.connect() as conn:
        backup_status.record_failure(
            conn,
            now_iso="2026-08-22T04:00:00+09:00",
            failure_summary=backup_status.FAILURE_EXECUTION,
        )
    with TestClient(main.app) as client:
        _session(client, email="admin@example.com", is_admin=True)
        response = client.get("/admin/settings")

    assert response.status_code == 200
    assert "최근 실행 실패" in response.text
    assert "마지막 시도 2026-08-22T04:00:00+09:00" in response.text
    assert backup_status.FAILURE_EXECUTION in response.text
    assert "환경변수 설정 여부가 아니라 실제 백업" in response.text


def test_admin_can_make_and_download_last_completed_weekly_xlsx(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    monkeypatch.setattr(
        dashboard_maintenance.clock,
        "now_kst",
        lambda: datetime.fromisoformat("2026-08-24T04:10:00+09:00"),
    )
    runtime._PIPELINE = DemoPipeline()
    with TestClient(main.app) as client:
        csrf = _session(client, email="admin@example.com", is_admin=True)
        made = client.post(
            "/admin/settings/weekly-reports/run",
            data={"csrf_token": csrf},
            follow_redirects=False,
        )
        downloaded = client.get("/admin/settings/weekly-reports/2026-08-17/download")

    assert made.status_code == 303
    assert downloaded.status_code == 200
    assert downloaded.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert load_workbook(BytesIO(downloaded.content)).sheetnames == [
        "한눈에 보기", "친구 이용", "피드백·문제"
    ]


def test_admin_button_and_internal_cron_use_the_same_maintenance_service(
    monkeypatch, tmp_path
):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    calls = []

    def shared_service(
        connect,
        *,
        operation,
        actor_email=dashboard_maintenance.SYSTEM_ACTOR_EMAIL,
        cleanup_runner=None,
    ):
        calls.append((connect, operation, actor_email, cleanup_runner))
        return dashboard_maintenance.MaintenanceResult(
            operation=operation,
            period_key=(
                "2026-08-17"
                if operation == dashboard_maintenance.OPERATION_WEEKLY
                else "2026-08-24"
            ),
            status="already_done",
        )

    monkeypatch.setattr(
        dashboard_maintenance,
        "run_current_operation",
        shared_service,
    )
    runtime._PIPELINE = DemoPipeline()
    with TestClient(main.app) as client:
        csrf = _session(client, email="admin@example.com", is_admin=True)
        manual_weekly = client.post(
            "/admin/settings/weekly-reports/run",
            data={"csrf_token": csrf},
            follow_redirects=False,
        )
        manual_cleanup = client.post(
            "/admin/settings/trash-cleanup/run",
            data={"csrf_token": csrf},
            follow_redirects=False,
        )
    cron_weekly = maintenance_router._run(dashboard_maintenance.OPERATION_WEEKLY)
    cron_cleanup = maintenance_router._run(dashboard_maintenance.OPERATION_CLEANUP)

    assert manual_weekly.status_code == 303
    assert manual_cleanup.status_code == 303
    assert cron_weekly.status == "already_done"
    assert cron_cleanup.status == "already_done"
    assert [call[1] for call in calls] == [
        dashboard_maintenance.OPERATION_WEEKLY,
        dashboard_maintenance.OPERATION_CLEANUP,
        dashboard_maintenance.OPERATION_WEEKLY,
        dashboard_maintenance.OPERATION_CLEANUP,
    ]
    assert calls[0][2] == "admin@example.com"
    assert calls[1][2] == "admin@example.com"
    assert calls[2][2] == dashboard_maintenance.SYSTEM_ACTOR_EMAIL
    assert calls[3][2] == dashboard_maintenance.SYSTEM_ACTOR_EMAIL
    assert calls[0][3] is None
    assert calls[1][3] is report_retention_adapter.purge_expired_reports
    assert calls[2][3] is None
    assert calls[3][3] is report_retention_adapter.purge_expired_reports
    assert cron_weekly.period_key == "2026-08-17"
    assert cron_cleanup.period_key == "2026-08-24"


def test_admin_trash_blocks_public_result_until_restore(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with TestClient(main.app) as client:
        csrf = _session(client, email="admin@example.com", is_admin=True)
        trashed = client.post(
            f"/admin/reports/{report_id}/trash",
            data={"reason": "duplicate", "csrf_token": csrf},
            follow_redirects=False,
        )
        blocked = client.get(f"/result/{report_id}", follow_redirects=False)
        notion_blocked = client.post(
            f"/notion/{report_id}",
            data={"csrf_token": csrf},
            follow_redirects=False,
        )
        restored = client.post(
            f"/admin/reports/{report_id}/restore",
            data={"reason": "keep", "csrf_token": csrf},
            follow_redirects=False,
        )

    assert trashed.status_code == 303
    assert blocked.status_code == 409
    assert notion_blocked.status_code == 409
    assert restored.status_code == 303


def test_corrected_snapshot_is_required_before_republish_and_overrides_live_memory(monkeypatch, tmp_path):
    monkeypatch.setenv(storage_constants.ENV_DB_PATH, str(tmp_path / "storage.db"))
    runtime._PIPELINE = DemoPipeline()
    report_id = _seed_report()
    with db.connect() as conn:
        original = reports.load(conn, report_id)
        assert original is not None
        dashboard_store.register_report(
            conn,
            report_id=report_id,
            corp_type=original.corp_type,
            payload_json=reports.report_to_json(original),
            now_iso="2026-08-22T10:00:00+09:00",
        )
        dashboard_store.record_error(
            conn,
            report_id=report_id,
            actor_email="member@example.com",
            area="출처",
            reason="수정이 필요합니다",
            now_iso="2026-08-22T10:01:00+09:00",
        )
        for status in (dashboard_store.REPORT_STATUS_FIXING, dashboard_store.REPORT_STATUS_RECHECKING):
            dashboard_store.change_report_state(
                conn,
                report_id=report_id,
                next_status=status,
                actor_email="admin@example.com",
                reason="수정본 재검사",
                now_iso="2026-08-22T10:02:00+09:00",
            )
        corrected_payload = reports.report_to_json(
            replace(original, generated_at="2026-08-20")
        )

    with TestClient(main.app) as client:
        csrf = _session(client, email="admin@example.com", is_admin=True)
        early_publish = client.post(
            f"/admin/reports/{report_id}/state",
            data={
                "status": "normal", "company_type": "listed",
                "reason": "수정본 수동 재공개", "csrf_token": csrf,
            },
            follow_redirects=False,
        )
        registered = client.post(
            f"/admin/reports/{report_id}/corrected-payload",
            data={"corrected_payload_json": corrected_payload, "csrf_token": csrf},
            follow_redirects=False,
        )
        published = client.post(
            f"/admin/reports/{report_id}/state",
            data={
                "status": "normal", "company_type": "listed",
                "reason": "수정본 수동 재공개", "csrf_token": csrf,
            },
            follow_redirects=False,
        )
        detail = client.get(f"/admin/reports/{report_id}")

    assert early_publish.status_code == 400
    assert registered.status_code == 303
    assert published.status_code == 303
    assert detail.status_code == 200
    assert "생성 기준 2026-08-20" in detail.text
    assert "현재 신고는 모두 처리되었습니다" in detail.text
    approved = reports_router._approved_public_report(report_id, original)
    assert approved is not None
    assert approved.generated_at == "2026-08-20"
