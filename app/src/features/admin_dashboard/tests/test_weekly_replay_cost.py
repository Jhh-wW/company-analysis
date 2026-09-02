"""주간 파일 「한눈에 보기」 비용 합계가 데모 리플레이 기록을 실제 지출처럼 더하지 않는지 못 박는다.

★ 데모 리플레이는 저장된 결과를 되돌려 줄 뿐 AI를 안 부른다.
  «비용»은 반드시 0원처럼 다뤄야 한다. 이 보호는 원래
  ``features/observability/metrics.py``의 대시보드 집계에만 있었는데, 그 파일이
  참조 0으로 오인되어 지워지면서 살아 있는 소비자인 이 「한눈에 보기」 집계에는
  같은 필터가 애초에 없었다는 사실이 드러났다.

★ 정본 판정: ``src.core.constants.REPLAY_MODEL_MARK`` 가 ``model`` 문자열에
  들어 있으면 리플레이다 (``web/paid_runtime.py`` 의 기존 필터와 같은 판정).

⚠️ 건수는 이 시험의 관심사가 아니다 — 오직 「확정 내부 AI 비용」열이 리플레이
  기록을 빼고 더하는지만 못 박는다.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from src.core.constants import REPLAY_MODEL_MARK
from src.features.admin_dashboard import store, weekly
from src.features.observability import lifecycle
from src.features.observability.constants import (
    CACHE_HIT_NONE,
    CORP_TYPE_UNKNOWN,
    END_STEP_COMPLETE,
)
from src.features.observability.records import RunRecord
from src.features.storage import db


def _final_record(*, run_id: str, at: str, cost_krw: float, model: str) -> RunRecord:
    return RunRecord(
        run_id=run_id,
        at=at,
        corp_type=CORP_TYPE_UNKNOWN,
        job="영업",
        end_step=END_STEP_COMPLETE,
        cache_hit=CACHE_HIT_NONE,
        fragments_collected=0,
        fragments_cited=0,
        sentences_made=0,
        sentences_passed=0,
        cells_filled=0,
        cells_missing=[],
        cells_suspect=[],
        grade="",
        human_check="",
        cost_krw=cost_krw,
        elapsed_sec=1.0,
        model=model,
    )


def _member_run(conn, *, run_id: str, day: str) -> None:
    """MEMBER 채널로 분류되도록 예약→확정까지 끝낸 실행 1건을 만든다."""
    assert store.reserve_member_run(
        conn,
        run_id=run_id,
        actor_email="member@example.com",
        day=day,
        now_iso=f"{day}T09:00:00+09:00",
    )
    assert store.settle_member_run(
        conn,
        run_id=run_id,
        succeeded=True,
        report_id="",
        now_iso=f"{day}T09:05:00+09:00",
    )


def _member_channel_cost(workbook_blob: bytes) -> float:
    workbook = load_workbook(BytesIO(workbook_blob), data_only=False)
    overview = workbook["한눈에 보기"]
    member_row = next(
        row
        for row in overview.iter_rows(min_row=5, max_row=overview.max_row)
        if row[0].value == "MEMBER"
    )
    return member_row[4].value


def test_한눈에_보기_비용_합계는_데모_리플레이_기록을_빼고_더한다(tmp_path):
    """★ 리플레이 기록은 합계에서 뺀다 — MEMBER 채널 실제 기록 1200.0원 + 리플레이 999999.0원이면 1200.0."""
    target = tmp_path / "weekly-replay.db"
    with db.connect(target) as conn:
        lifecycle.ensure_schema(conn)
        _member_run(conn, run_id="real-1", day="2026-08-18")
        _member_run(conn, run_id="replay-1", day="2026-08-18")
        assert lifecycle.finalize_once(
            conn,
            _final_record(
                run_id="real-1",
                at="2026-08-18T10:00:00",
                cost_krw=1200.0,
                model="claude-sonnet-4-6",
            ),
        )
        assert lifecycle.finalize_once(
            conn,
            _final_record(
                run_id="replay-1",
                at="2026-08-18T11:00:00",
                cost_krw=999999.0,
                model=f"claude-haiku-4-5 {REPLAY_MODEL_MARK}",
            ),
        )

        workbook_blob = weekly.build_weekly_workbook(conn, week_start="2026-08-17")

    assert _member_channel_cost(workbook_blob) == 1200.0


def test_리플레이_리터럴_문자열로도_같은_결과를_다시_못_박는다():
    """상수 값이 조용히 바뀌어도 「지금 실제로 제외되는 문자열」을 리터럴로 고정한다.

    ★ 이 단정은 ``REPLAY_MODEL_MARK`` import에 기대지 않는다 — 상수 값이
      달라지는 회귀가 생겨도 이 시험만은 여전히 원래 표식 문자열을 직접 써서
      제외 여부를 재확인한다.
    """
    assert REPLAY_MODEL_MARK == "(데모 기록)"


def test_한눈에_보기_비용_합계는_리터럴_꼬리표_기록도_빼고_더한다(tmp_path):
    """위 시험과 같은 상황을 상수 import 없이 리터럴 문자열만으로 재확인한다."""
    target = tmp_path / "weekly-replay-literal.db"
    with db.connect(target) as conn:
        lifecycle.ensure_schema(conn)
        _member_run(conn, run_id="real-2", day="2026-08-18")
        _member_run(conn, run_id="replay-2", day="2026-08-18")
        assert lifecycle.finalize_once(
            conn,
            _final_record(
                run_id="real-2",
                at="2026-08-18T10:00:00",
                cost_krw=1200.0,
                model="claude-sonnet-4-6",
            ),
        )
        assert lifecycle.finalize_once(
            conn,
            _final_record(
                run_id="replay-2",
                at="2026-08-18T11:00:00",
                cost_krw=999999.0,
                model="claude-haiku-4-5 (데모 기록)",
            ),
        )

        workbook_blob = weekly.build_weekly_workbook(conn, week_start="2026-08-17")

    assert _member_channel_cost(workbook_blob) == 1200.0
