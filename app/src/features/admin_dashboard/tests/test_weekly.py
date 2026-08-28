from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from src.features.admin_dashboard import store, weekly
from src.features.pipeline.canonical_demo import build_demo_report
from src.features.storage import db, reports


def test_empty_weekly_workbook_has_the_three_required_sheets(tmp_path):
    target = tmp_path / "weekly.db"
    with db.connect(target) as conn:
        workbook_blob = weekly.build_weekly_workbook(conn, week_start="2026-08-17")

    workbook = load_workbook(BytesIO(workbook_blob), data_only=False)

    assert workbook.sheetnames == ["한눈에 보기", "친구 이용", "피드백·문제"]
    assert workbook["한눈에 보기"].freeze_panes == "A5"
    assert workbook["친구 이용"].freeze_panes == "A5"
    assert workbook["피드백·문제"].freeze_panes == "A5"


def test_xlsx_value_boundary_preserves_types_and_neutralizes_formula_prefixes():
    assert weekly._safe_xlsx_value(7) == 7
    assert weekly._safe_xlsx_value(12.5) == 12.5
    assert weekly._safe_xlsx_value("일반 문장") == "일반 문장"
    assert weekly._safe_xlsx_value("=1+1") == "'=1+1"
    assert weekly._safe_xlsx_value("  +SUM(A1:A2)") == "'  +SUM(A1:A2)"
    assert weekly._safe_xlsx_value("\u200b-10") == "'\u200b-10"
    assert weekly._safe_xlsx_value("\ufeff@SUM(A1:A2)") == "'\ufeff@SUM(A1:A2)"
    assert weekly._safe_xlsx_value("\x01=1+1") == "'�=1+1"


def test_member_text_is_never_an_executable_formula_in_weekly_xlsx(tmp_path):
    target = tmp_path / "weekly-formula-boundary.db"
    with db.connect(target) as conn:
        store.save_survey(
            conn,
            report_id="report-formula-boundary",
            actor_email="member@example.com",
            rating=5,
            overall_feedback='=HYPERLINK("https://attacker.invalid", "열기")',
            business_distinction="+SUM(A1:A2)",
            add_information="\u200b-10+20",
            delete_information="@SUM(A1:A2)",
            now_iso="2026-08-18T10:00:00+09:00",
        )
        store.record_error(
            conn,
            report_id="report-formula-boundary",
            actor_email="member@example.com",
            area="=WEBSERVICE(\"https://attacker.invalid\")",
            reason="\ufeff+CMD",
            now_iso="2026-08-18T10:01:00+09:00",
        )
        workbook_blob = weekly.build_weekly_workbook(
            conn,
            week_start="2026-08-17",
        )

    workbook = load_workbook(BytesIO(workbook_blob), data_only=False)
    sheet = workbook["피드백·문제"]

    assert sheet["B5"].value == 5
    assert sheet["B5"].data_type == "n"
    assert sheet["E5"].value.startswith("'=")
    assert sheet["F5"].value.startswith("'+")
    assert sheet["G5"].value.startswith("'\u200b-")
    assert sheet["H5"].value.startswith("'@")
    assert sheet["C9"].value.startswith("'=")
    assert sheet["D9"].value.startswith("'\ufeff+")
    assert all(
        cell.data_type != "f"
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )


def test_weekly_file_is_saved_once_for_a_week_and_downloadable_from_storage(tmp_path):
    target = tmp_path / "weekly.db"
    with db.connect(target) as conn:
        key = store.claim_operation(
            conn,
            operation=store.OPERATION_WEEKLY_XLSX,
            period_key="2026-08-17",
            actor_email="admin@example.com",
            now_iso="2026-08-24T04:10:00+09:00",
        )
        assert key == "weekly_xlsx:2026-08-17"
        workbook_blob = weekly.build_weekly_workbook(conn, week_start="2026-08-17")
        assert store.save_weekly_report(
            conn,
            week_start="2026-08-17",
            workbook_blob=workbook_blob,
            actor_email="admin@example.com",
            now_iso="2026-08-24T04:10:01+09:00",
        )
        assert store.complete_operation(
            conn,
            key=key,
            status="succeeded",
            detail="manual weekly workbook",
            now_iso="2026-08-24T04:10:01+09:00",
        )
        assert store.claim_operation(
            conn,
            operation=store.OPERATION_WEEKLY_XLSX,
            period_key="2026-08-17",
            actor_email="admin@example.com",
            now_iso="2026-08-24T04:11:00+09:00",
        ) is None
        assert store.load_weekly_report_blob(conn, week_start="2026-08-17") == workbook_blob


def test_cleanup_closes_only_previous_day_stalled_operations_as_failed(tmp_path):
    target = tmp_path / "weekly.db"
    with db.connect(target) as conn:
        key = store.claim_operation(
            conn,
            operation=store.OPERATION_WEEKLY_XLSX,
            period_key="2026-08-17",
            actor_email="admin@example.com",
            now_iso="2026-08-17T04:10:00+09:00",
        )
        assert key is not None
        assert store.fail_stalled_operations(
            conn,
            before_iso="2026-08-18T00:00:00+09:00",
            now_iso="2026-08-18T04:20:00+09:00",
        ) == 1
        assert store.fail_stalled_operations(
            conn,
            before_iso="2026-08-18T00:00:00+09:00",
            now_iso="2026-08-18T04:20:01+09:00",
        ) == 0
        claims = store.list_operation_claims(conn)
        issues = store.list_failed_operation_issues(conn)

    assert claims[0]["status"] == "failed"
    assert issues[0]["detail"] == "previous_kst_day_not_finished"


def test_trash_immediately_excludes_statistics_and_purges_report_after_30_days(tmp_path):
    target = tmp_path / "weekly.db"
    report_id = "trash-report"
    with db.connect(target) as conn:
        report = build_demo_report()
        reports.save(
            conn,
            report_id,
            "CORP-001",
            "",
            report,
            created_at="2026-07-01T09:00:00+09:00",
        )
        store.register_report(
            conn,
            report_id=report_id,
            corp_type=report.corp_type,
            payload_json=reports.report_to_json(report),
            now_iso="2026-07-01T09:00:00+09:00",
        )
        store.save_survey(
            conn,
            report_id=report_id,
            actor_email="member@example.com",
            rating=5,
            overall_feedback="good",
            business_distinction="clear",
            add_information="",
            delete_information="",
            now_iso="2026-07-01T10:00:00+09:00",
        )
        store.record_error(
            conn,
            report_id=report_id,
            actor_email="member@example.com",
            area="source",
            reason="mismatch",
            now_iso="2026-07-01T10:01:00+09:00",
        )
        assert store.reserve_member_run(
            conn,
            run_id="trash-run",
            actor_email="member@example.com",
            day="2026-07-01",
            now_iso="2026-07-01T09:00:00+09:00",
        )
        assert store.settle_member_run(
            conn,
            run_id="trash-run",
            succeeded=True,
            report_id=report_id,
            outcome="report",
            company_type=store.COMPANY_LISTED,
            cost_krw=100,
            now_iso="2026-07-01T10:02:00+09:00",
        )
        assert store.survey_summary(conn) == (1, 1)
        assert len(store.list_open_errors(conn)) == 1
        assert store.member_run_statistics(conn)["settled"]["used"] == 1

        assert store.trash_report(
            conn,
            report_id=report_id,
            actor_email="admin@example.com",
            reason="duplicate",
            now_iso="2026-07-01T11:00:00+09:00",
        )
        assert store.report_is_trashed(conn, report_id)
        assert store.survey_summary(conn) == (0, 0)
        assert store.list_open_errors(conn) == []
        assert store.member_run_statistics(conn)["settled"]["used"] == 0

        assert store.purge_expired_trash(conn, now_iso="2026-07-31T11:00:00+09:00") == 1
        assert reports.load(conn, report_id) is None
        assert store.trash_record(conn, report_id).status == store.TRASH_PURGED
