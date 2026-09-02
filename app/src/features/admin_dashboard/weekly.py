"""관리자 전용 주간 운영 파일을 기존 SQLite 기록만으로 만든다.

새 조사·AI 호출·외부 연결은 하지 않는다. 각 이용 경로를 합산하지 않고 분리해
보여 주며, 기존 이력에 없는 값은 ``확인 불가``로 남긴다.
"""

from __future__ import annotations

from io import BytesIO
import re
from typing import Any
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.constants import REPLAY_MODEL_MARK
from src.features.admin_dashboard import store
from src.features.observability import constants as observability_constants
from src.features.observability import lifecycle
from src.features.sharelink import store as share_store
from src.features.storage import reports as report_store


_NAVY = "17365D"
_BLUE = "DCE6F1"
_PALE = "F6F8FB"
_WHITE = "FFFFFF"
_THIN = Side(style="thin", color="B8C2CC")
_EXCEL_FORMULA_PREFIXES = frozenset("=+-@")
_ILLEGAL_XLSX_CONTROLS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
_COMPANY_LABELS = {
    store.COMPANY_LISTED: "상장사",
    store.COMPANY_AUDITED: "비상장 외감",
    store.COMPANY_UNDECIDED: "판정 전 종료",
}


def _safe_xlsx_value(value: Any) -> Any:
    """외부 문자열이 관리자 PC에서 Excel 명령으로 실행되지 않게 한다.

    XLSX의 문자열은 CSV가 아니어도 ``=``로 시작하면 수식으로 저장될 수 있다.
    Excel이 앞에서 무시할 수 있는 공백·제어·서식 문자 뒤의 식별 문자까지 보고
    텍스트 표식인 작은따옴표를 붙인다. 숫자 같은 비문자 값은 서식을 잃지 않게
    그대로 둔다. XML에 넣을 수 없는 제어 문자는 파일 전체 생성을 깨지 않도록
    대체 문자로 바꾼다.
    """

    if not isinstance(value, str):
        return value
    first_visible = next(
        (
            character
            for character in value
            if not character.isspace()
            and not unicodedata.category(character).startswith("C")
        ),
        "",
    )
    safe_text = _ILLEGAL_XLSX_CONTROLS.sub("�", value)
    if first_visible in _EXCEL_FORMULA_PREFIXES:
        return "'" + safe_text
    return safe_text


def _write_cell(sheet, row: int, column: int, value: Any):
    """주간 파일의 모든 셀 값을 단 하나의 출력 경계로 쓴다."""

    return sheet.cell(row, column, _safe_xlsx_value(value))


def build_weekly_workbook(conn, *, week_start: str) -> bytes:
    """지난 월~일의 세 시트를 갖는 XLSX 바이트를 만든다."""
    lifecycle.ensure_schema(conn)
    start, end = store.weekly_period(week_start)
    data = _collect_weekly_data(conn, start=start, end=end)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "한눈에 보기"
    _write_overview(overview, start=start, end=end, rows=data["overview"])

    member_sheet = workbook.create_sheet("친구 이용")
    _write_table_sheet(
        member_sheet,
        title="친구 이용",
        subtitle=f"기간: {start} ~ {end} · MEMBER 기록만 표시합니다.",
        headers=("기준일", "친구", "회사", "회사 유형", "결과", "저장 보고서", "저장 보고서 재사용 여부", "내부 AI 비용", "비용 상태"),
        rows=data["members"],
        widths=(14, 30, 28, 16, 14, 24, 28, 16, 14),
        number_columns={8: "#,##0"},
    )

    feedback_sheet = workbook.create_sheet("피드백·문제")
    _write_feedback_sheet(
        feedback_sheet,
        start=start,
        end=end,
        feedback_rows=data["feedback"],
        issue_rows=data["issues"],
    )

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _collect_weekly_data(conn, *, start: str, end: str) -> dict[str, list[tuple[Any, ...]]]:
    trashed_run_ids = _trashed_run_ids(conn)
    final_records = [
        record for record in lifecycle.list_final(conn)
        if start <= record.at[:10] <= end and record.run_id not in trashed_run_ids
    ]
    member_rows_by_run = _latest_member_rows(conn, start=start, end=end)
    link_run_ids = _link_run_ids(conn)

    by_channel: dict[str, list] = {"MEMBER": [], "LINK": [], "확인 불가": []}
    for record in final_records:
        channel = "MEMBER" if record.run_id in member_rows_by_run else (
            "LINK" if record.run_id in link_run_ids else "확인 불가"
        )
        by_channel[channel].append(record)

    overview = []
    for channel in ("MEMBER", "LINK", "확인 불가"):
        records = by_channel[channel]
        successes = sum(
            record.end_step == observability_constants.END_STEP_COMPLETE for record in records
        )
        # ★ 비용만 데모 리플레이 기록을 뺀다 (문제로그 P-84). 데모는 저장된 결과를
        #   되돌려 줄 뿐 AI를 안 부르므로 0원처럼 다뤄야 한다. 건수(len(records)·
        #   successes)는 그대로 센다 — 데모도 «실행»은 실행이다. 지워졌던
        #   observability/metrics.py의 build_dashboard와 같은 판정·같은 결정이다.
        costs = sum(
            record.cost_krw for record in records
            if REPLAY_MODEL_MARK not in (record.model or "")
        )
        overview.append((channel, len(records), successes, len(records) - successes, costs, "0"))

    new_issues = _issue_rows(conn, start=start, end=end)
    feedback_rows = _feedback_rows(conn, start=start, end=end)
    open_issues = sum(
        not store.report_is_trashed(conn, error.report_id)
        for error in store.list_open_errors(conn, limit=500)
    )
    helpful = sum(int(row[1]) >= 4 for row in feedback_rows)
    satisfaction = (
        "자료 모으는 중"
        if len(feedback_rows) < 5
        else f"{round(helpful * 100 / len(feedback_rows))}% ({helpful}/{len(feedback_rows)})"
    )
    overview.extend(
        (
            ("새 문제", len(new_issues), "", "", "", ""),
            ("미해결 문제", open_issues, "", "", "", ""),
            ("설문", len(feedback_rows), "", "", "", satisfaction),
        )
    )

    members: list[tuple[Any, ...]] = []
    for row in member_rows_by_run.values():
        report_id = str(row["report_id"])
        report = None if store.report_is_trashed(conn, report_id) else report_store.load(conn, report_id)
        members.append(
            (
                str(row["day"]),
                str(row["actor_email"]),
                report.company if report is not None else "확인 불가",
                _COMPANY_LABELS.get(str(row["company_type"]), "판정 전 종료"),
                "성공" if str(row["state"]) == store.MEMBER_USAGE_USED else "실패·반환",
                report_id or "—",
                "확인 불가 (재사용 이력 미저장)",
                float(row["cost_krw"]),
                "확정" if str(row["cost_state"]) == "confirmed" else (
                    "미확정" if str(row["cost_state"]) == "uncertain" else "발생 없음"
                ),
            )
        )
    members.sort(key=lambda row: (str(row[0]), str(row[1]), str(row[5])))
    return {
        "overview": overview,
        "members": members,
        "feedback": feedback_rows,
        "issues": new_issues,
    }


def _latest_member_rows(conn, *, start: str, end: str) -> dict[str, Any]:
    rows = conn.execute(
        f"""SELECT e.run_id, e.actor_email, e.day, e.state, e.company_type, e.report_id,
                   e.cost_krw, e.cost_state
        FROM {store.TABLE_MEMBER_RUN_SUMMARY_EVENTS} AS e
        JOIN (
            SELECT run_id, MAX(id) AS latest_id
            FROM {store.TABLE_MEMBER_RUN_SUMMARY_EVENTS}
            WHERE day >= ? AND day <= ?
            GROUP BY run_id
        ) AS latest ON latest.latest_id = e.id
        LEFT JOIN {store.TABLE_REPORT_TRASH} AS t ON t.report_id = e.report_id
        WHERE e.state IN (?, ?)
          AND (e.report_id = '' OR t.status IS NULL OR t.status = 'active')""",
        (start, end, store.MEMBER_USAGE_USED, store.MEMBER_USAGE_RETURNED),
    ).fetchall()
    return {str(row["run_id"]): row for row in rows}


def _trashed_run_ids(conn) -> set[str]:
    """휴지통 보고서와 결속된 실행은 주간 집계의 모든 시트에서 빼낸다."""
    member_rows = conn.execute(
        f"""SELECT DISTINCT e.run_id
        FROM {store.TABLE_MEMBER_RUN_SUMMARY_EVENTS} AS e
        JOIN {store.TABLE_REPORT_TRASH} AS t ON t.report_id = e.report_id
        WHERE e.report_id <> '' AND t.status IN ('trashed', 'purged')"""
    ).fetchall()
    link_rows = conn.execute(
        f"""SELECT DISTINCT h.run_id
        FROM {share_store.TABLE_RUN_HISTORY} AS h
        JOIN {store.TABLE_REPORT_TRASH} AS t ON t.report_id = h.report_id
        WHERE h.report_id <> '' AND t.status IN ('trashed', 'purged')"""
    ).fetchall()
    return {str(row["run_id"]) for row in (*member_rows, *link_rows)}


def _link_run_ids(conn) -> set[str]:
    rows = conn.execute(
        f"""SELECT h.run_id FROM {share_store.TABLE_RUN_HISTORY} AS h
        LEFT JOIN {store.TABLE_REPORT_TRASH} AS t ON t.report_id = h.report_id
        WHERE h.report_id = '' OR t.status IS NULL OR t.status = 'active'"""
    ).fetchall()
    return {str(row["run_id"]) for row in rows}


def _feedback_rows(conn, *, start: str, end: str) -> list[tuple[Any, ...]]:
    rows = conn.execute(
        f"""SELECT report_id, actor_email, rating, overall_feedback, business_distinction,
                   add_information, delete_information, created_at
        FROM {store.TABLE_SURVEY_EVENTS}
        WHERE substr(created_at, 1, 10) >= ? AND substr(created_at, 1, 10) <= ?
        ORDER BY created_at ASC, id ASC""",
        (start, end),
    ).fetchall()
    return [
        (
            str(row["created_at"]), int(row["rating"]), str(row["actor_email"]),
            str(row["report_id"]), str(row["overall_feedback"]),
            str(row["business_distinction"]), str(row["add_information"]),
            str(row["delete_information"]),
        )
        for row in rows
        if not store.report_is_trashed(conn, str(row["report_id"]))
    ]


def _issue_rows(conn, *, start: str, end: str) -> list[tuple[Any, ...]]:
    rows = conn.execute(
        f"""SELECT e.report_id, e.area, e.reason, e.created_at,
                   COALESCE(s.status, e.status) AS current_status
        FROM {store.TABLE_ERRORS} AS e
        LEFT JOIN {store.TABLE_REPORT_STATES} AS s ON s.report_id = e.report_id
        WHERE substr(e.created_at, 1, 10) >= ? AND substr(e.created_at, 1, 10) <= ?
        ORDER BY e.created_at ASC, e.id ASC""",
        (start, end),
    ).fetchall()
    labels = {
        store.REPORT_STATUS_PENDING: "확인 대기",
        store.REPORT_STATUS_FIXING: "수정 중",
        store.REPORT_STATUS_RECHECKING: "재검사 중",
        store.REPORT_STATUS_RECHECK_FAILED: "재검사 실패",
        store.REPORT_STATUS_NORMAL: "정상",
    }
    return [
        (
            str(row["created_at"]), str(row["report_id"]), str(row["area"]),
            str(row["reason"]), labels.get(str(row["current_status"]), "확인 불가"),
        )
        for row in rows
        if not store.report_is_trashed(conn, str(row["report_id"]))
    ]


def _write_overview(sheet, *, start: str, end: str, rows: list[tuple[Any, ...]]) -> None:
    _write_title(sheet, "한눈에 보기", f"기간: {start} ~ {end} · 이용 경로를 합산하지 않습니다.", 6)
    headers = ("이용 경로·항목", "생성", "성공", "실패", "확정 내부 AI 비용", "비고")
    _write_header(sheet, row=4, headers=headers)
    for index, row in enumerate(rows, start=5):
        for column, value in enumerate(row, start=1):
            _write_cell(sheet, index, column, value)
        sheet.cell(index, 5).number_format = "#,##0"
    _finish_sheet(sheet, max_row=max(5, 4 + len(rows)), max_col=6, widths=(20, 12, 12, 12, 20, 24))


def _write_feedback_sheet(sheet, *, start: str, end: str, feedback_rows, issue_rows) -> None:
    _write_title(sheet, "피드백·문제", f"기간: {start} ~ {end} · MEMBER 설문과 오류 신고를 분리해 기록합니다.", 8)
    feedback_headers = ("기록 시각", "별점", "친구", "보고서", "전체 소감", "사업상 구분점", "추가 희망", "삭제 희망")
    _write_header(sheet, row=4, headers=feedback_headers)
    row_number = 5
    for row in feedback_rows:
        for column, value in enumerate(row, start=1):
            _write_cell(sheet, row_number, column, value)
        row_number += 1
    if not feedback_rows:
        _write_cell(sheet, row_number, 1, "이 기간의 설문이 없습니다.")
        row_number += 1
    row_number += 1
    sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=5)
    section = _write_cell(sheet, row_number, 1, "오류 신고")
    section.font = Font(bold=True, color=_WHITE)
    section.fill = PatternFill("solid", fgColor=_NAVY)
    row_number += 1
    issue_headers = ("기록 시각", "보고서", "어느 부분", "이유", "처리 상태")
    _write_header(sheet, row=row_number, headers=issue_headers)
    row_number += 1
    for row in issue_rows:
        for column, value in enumerate(row, start=1):
            _write_cell(sheet, row_number, column, value)
        row_number += 1
    if not issue_rows:
        _write_cell(sheet, row_number, 1, "이 기간의 오류 신고가 없습니다.")
        row_number += 1
    _finish_sheet(sheet, max_row=row_number, max_col=8, widths=(20, 10, 28, 24, 42, 42, 34, 34))
    sheet.freeze_panes = "A5"


def _write_table_sheet(sheet, *, title: str, subtitle: str, headers, rows, widths, number_columns: dict[int, str]) -> None:
    _write_title(sheet, title, subtitle, len(headers))
    _write_header(sheet, row=4, headers=headers)
    for row_number, row in enumerate(rows, start=5):
        for column, value in enumerate(row, start=1):
            cell = _write_cell(sheet, row_number, column, value)
            if column in number_columns:
                cell.number_format = number_columns[column]
    if not rows:
        _write_cell(sheet, 5, 1, "이 기간의 기록이 없습니다.")
    _finish_sheet(sheet, max_row=max(5, 4 + len(rows)), max_col=len(headers), widths=widths)


def _write_title(sheet, title: str, subtitle: str, columns: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    title_cell = _write_cell(sheet, 1, 1, title)
    title_cell.font = Font(size=16, bold=True, color=_WHITE)
    title_cell.fill = PatternFill("solid", fgColor=_NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 26
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    subtitle_cell = _write_cell(sheet, 2, 1, subtitle)
    subtitle_cell.font = Font(italic=True, color="4A5568")
    subtitle_cell.fill = PatternFill("solid", fgColor=_PALE)
    subtitle_cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 32


def _write_header(sheet, *, row: int, headers) -> None:
    for column, header in enumerate(headers, start=1):
        cell = _write_cell(sheet, row, column, header)
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
    sheet.row_dimensions[row].height = 28


def _finish_sheet(sheet, *, max_row: int, max_col: int, widths) -> None:
    for row in sheet.iter_rows(min_row=4, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
            if cell.row % 2 == 1 and cell.row > 4:
                cell.fill = PatternFill("solid", fgColor="FBFCFE")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(float(width), 10), 45)
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(max_col)}{max(4, max_row)}"
    sheet.sheet_view.showGridLines = False
